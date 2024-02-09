import openpyxl


class ExcelManager:

    @staticmethod
    def OpenBook(bookName: str) -> openpyxl.Workbook:
        book = openpyxl.open(f"Excels/{bookName}.xlsx")
        return book

    @staticmethod
    def CreateBook(bookName: str) -> openpyxl.Workbook:
        book = openpyxl.Workbook()
        book.save(f"Excels/{bookName}.xlsx")
        return book

    @staticmethod
    def write_to_excel(file_path: str, values, column_num = 1):
        # column_num - Номер столбца, в который будем записывать значения
        # values - Массив значений
        # file_path - Имя файла

        # Открываем файл Excel
        wb = openpyxl.load_workbook(file_path)

        # Выбираем активный лист
        sheet = wb.active

        # Проходим по значениям и записываем их во второй столбец
        for row_num, value in enumerate(values, start=1):
            # Записываем значение в указанную ячейку
            sheet.cell(row=row_num, column=column_num).value = value

        # Сохраняем изменения в файл
        wb.save(file_path)
        wb.close()

ExcelManager.write_to_excel('conn.xlsx', (1, 2, 3, 4, 5), 7)